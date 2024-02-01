from openpyxl import Workbook, load_workbook

wb= load_workbook(filename='Sheet.xlsx')
sh= wb['name']

#print(sh['A2'].value)
row_ct= sh.max_row
col_ct= sh.max_column
for i in range (1, row_ct+1):
    for j in range(1, col_ct + 1):
        print(sh.cell(row=i, column=j).value, end= ' ')
    print('\n')
